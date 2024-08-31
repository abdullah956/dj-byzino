from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.core.mail import send_mail
from django.conf import settings
from django.db.models import Sum
from io import BytesIO
from openpyxl import Workbook
import openpyxl
from categories.models import Product
from orders.models import Order
from users.models import ContactMessage, Subscriber
from django.contrib.auth import get_user_model
from django.contrib import messages

#dashboard
def stats_view(request):
    total_subscribers = Subscriber.objects.count()
    total_orders = Order.objects.count()
    total_products = Product.objects.count()
    total_earnings = Order.objects.filter(status='delivered').aggregate(total_amount=Sum('amount'))['total_amount'] or 0
    order_list = Order.objects.all().order_by('-order_date')
    paginator = Paginator(order_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'total_subscribers': total_subscribers,
        'total_orders': total_orders,
        'total_products': total_products,
        'total_earnings': total_earnings,
        'recent_orders': page_obj,
    }
    return render(request, 'dashboard/stats.html', context)


#order detail
def order_detail_view(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    product_data = order.products
    product_ids = [item['product_id'] for item in product_data]
    products = Product.objects.filter(id__in=product_ids)
    products_with_quantities = []
    for item in product_data:
        product_id = item['product_id']
        quantity = item['quantity']
        product = products.filter(id=product_id).first()
        if product:
            products_with_quantities.append({
                'product': product,
                'quantity': quantity
            })
    context = {
        'order': order,
        'products_with_quantities': products_with_quantities,
    }
    return render(request, 'dashboard/order_detail.html', context)


#messages  
def message_list(request):
    messages = ContactMessage.objects.all().order_by('-created_at')
    paginator = Paginator(messages, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'dashboard/message_list.html', {'page_obj': page_obj})

#message show
def message_detail(request, pk):
    message = get_object_or_404(ContactMessage, pk=pk)
    if not message.is_read:
        message.is_read = True
        message.save()

    return render(request, 'dashboard/message_detail.html', {'message': message})


#all users
User = get_user_model()
def user_list_view(request):
    users = User.objects.all()
    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    if request.GET.get('export') == 'excel':
        email_list = [user.email for user in users]
        wb = Workbook()
        ws = wb.active
        ws.title = 'Users'
        ws.append(['Email'])
        for email in email_list:
            ws.append([email])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=emails.xlsx'
        stream = BytesIO()
        wb.save(stream)
        response.write(stream.getvalue())
        return response

    return render(request, 'dashboard/user_list.html', {'page_obj': page_obj})

#export excel users
def export_users_to_excel(request):
    users = User.objects.all()
    email_list = [user.email for user in users]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Users'
    ws.append(['Email'])
    for email in email_list:
        ws.append([email])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="users_emails.xlsx"'
    return response

#subs lsit
def subscriber_list(request):
    subscribers = Subscriber.objects.all()
    paginator = Paginator(subscribers, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    print(subscribers)
    return render(request, 'dashboard/subscriber_list.html', {'subscribers': page_obj})

#export subs excel
def export_subscribers_to_excel(request):
    subscribers = Subscriber.objects.all()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Subscribers'
    headers = ['Email']
    worksheet.append(headers)
    for subscriber in subscribers:
        worksheet.append([
            subscriber.email,
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=subscribers.xlsx'
    workbook.save(response)
    return response

#manage orders
def manage_orders(request):
    if not request.user.is_staff:
        return redirect('login')

    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        order = Order.objects.get(id=order_id)

        new_status = request.POST.get('status')
        if new_status != order.status:
            order.status = new_status
            order.save()

            if new_status == 'canceled':
                send_mail(
                    'Order Canceled',
                    f'Your order #{order.id} has been canceled.',
                    settings.DEFAULT_FROM_EMAIL,
                    [order.email],
                )
                messages.success(request, f'Order #{order.id} status updated to "Canceled".')
            elif new_status == 'delivered':
                send_mail(
                    'Order Delivered',
                    f'Your order #{order.id} has been delivered.',
                    settings.DEFAULT_FROM_EMAIL,
                    [order.email],
                )
                messages.success(request, f'Order #{order.id} status updated to "Delivered".')

        new_is_shipped = 'is_shipped' in request.POST
        if new_is_shipped != order.is_shipped:
            order.is_shipped = new_is_shipped
            order.save()

            if new_is_shipped:
                send_mail(
                    'Order Shipped',
                    f'Your order #{order.id} has been shipped.',
                    settings.DEFAULT_FROM_EMAIL,
                    [order.email],
                )
                messages.success(request, f'Order #{order.id} marked as shipped.')

        new_is_paid = 'is_paid' in request.POST
        if new_is_paid != order.is_paid:
            order.is_paid = new_is_paid
            order.save()

            if new_is_paid:
                messages.success(request, f'Order #{order.id} marked as paid.')
            else:
                messages.success(request, f'Order #{order.id} marked as unpaid.')

    orders = Order.objects.all().order_by('-order_date')
    paginator = Paginator(orders, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'dashboard/manage_orders.html', {'orders': page_obj})
